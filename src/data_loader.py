"""
data_loader.py — Prayer times data management.

Handles fetching prayer times from the abdus.dev Diyanet API,
caching them locally, and loading from Excel as fallback.
"""

import os
import json
import logging
from datetime import datetime, date, timedelta
from typing import Optional

logger = logging.getLogger(__name__)

# abdus.dev Diyanet API wrapper
API_BASE = "https://prayertimes.api.abdus.dev/api/diyanet"


def search_locations(query: str) -> list[dict]:
    """Search for locations by name via the Diyanet API.
    
    Args:
        query: City or region name to search for.
        
    Returns:
        List of location dicts: [{'id': int, 'country': str, 'city': str, 'region': str}, ...]
    """
    import urllib.request
    import urllib.parse
    
    url = f"{API_BASE}/search?q={urllib.parse.quote(query)}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'PrayerWidget/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            return data if isinstance(data, list) else []
    except Exception as e:
        logger.error(f"Location search failed: {e}")
        return []


def fetch_prayer_times(location_id: int) -> list[dict]:
    """Fetch prayer times from the Diyanet API for a location.
    
    The API returns approximately 30 days of data from today.
    
    Args:
        location_id: Diyanet location ID (e.g. 9206 for Ankara).
        
    Returns:
        List of prayer time dicts with keys: date, fajr, sun, dhuhr, asr, maghrib, isha
    """
    import urllib.request
    
    url = f"{API_BASE}/prayertimes?location_id={location_id}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'PrayerWidget/1.0'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            return data if isinstance(data, list) else []
    except Exception as e:
        logger.error(f"Prayer times fetch failed: {e}")
        return []


def save_cache(cache_path: str, location_id: int, prayer_data: list[dict]) -> None:
    """Save fetched prayer times to local JSON cache.
    
    Args:
        cache_path: Path to the cache JSON file.
        location_id: The location ID this data belongs to.
        prayer_data: Raw API response list.
    """
    cache = {
        'location_id': location_id,
        'fetched_at': datetime.now().isoformat(),
        'data': prayer_data,
    }
    try:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
        logger.info(f"Cache saved: {len(prayer_data)} days to {cache_path}")
    except Exception as e:
        logger.error(f"Cache save failed: {e}")


def load_cache(cache_path: str) -> Optional[dict]:
    """Load cached prayer times from disk.
    
    Returns:
        Cache dict with 'location_id', 'fetched_at', 'data' keys, or None if not found.
    """
    try:
        if os.path.exists(cache_path):
            with open(cache_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Cache load failed: {e}")
    return None


def parse_api_data(prayer_data: list[dict]) -> dict:
    """Parse API response into the internal format.
    
    Converts the API response into a dict mapping date objects to prayer time dicts.
    The internal format uses canonical keys: fajr, sunrise, dhuhr, asr, maghrib, isha.
    
    Args:
        prayer_data: Raw API response list.
        
    Returns:
        Dict mapping date -> {'fajr': 'HH:MM', 'sunrise': 'HH:MM', ...}
    """
    daily_times = {}
    for entry in prayer_data:
        try:
            # Parse date: "2026-05-24T00:00:00"
            date_str = entry.get('date', '')
            if 'T' in date_str:
                date_str = date_str.split('T')[0]
            d = date.fromisoformat(date_str)
            
            # Map API keys to internal canonical keys
            daily_times[d] = {
                'fajr': entry.get('fajr', '').strip(),
                'sunrise': entry.get('sun', '').strip(),
                'dhuhr': entry.get('dhuhr', '').strip(),
                'asr': entry.get('asr', '').strip(),
                'maghrib': entry.get('maghrib', '').strip(),
                'isha': entry.get('isha', '').strip(),
            }
        except (ValueError, AttributeError) as e:
            logger.warning(f"Skipping invalid entry: {entry} — {e}")
            continue
    
    return daily_times


def load_excel_data(excel_path: str) -> tuple[dict, dict, str]:
    """Load prayer times and hijri dates from an Excel file.
    
    The Excel format must match the Diyanet export:
    - City name in cell B1
    - Headers in row 2
    - Data from row 3 onwards
    - Columns: Date | Weekday | Fajr | Sunrise | Dhuhr | Asr | Maghrib | Isha | ... | Hijri Date
    
    Args:
        excel_path: Path to the .xlsx file.
        
    Returns:
        Tuple of (daily_prayer_times, hijri_dates, city_name)
        - daily_prayer_times: dict mapping date -> prayer times dict
        - hijri_dates: dict mapping date -> hijri date string
        - city_name: str from the Excel header
    """
    import openpyxl
    
    daily_times = {}
    hijri_dates = {}
    city_name = "Prayer Times"
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        # City name from B1
        raw_city = sheet.cell(row=1, column=2).value
        if raw_city:
            city_name = str(raw_city).strip().title()
        
        # Parse rows (skip header rows 1-2)
        for row in sheet.iter_rows(min_row=3, values_only=False):
            cells = [c.value for c in row]
            if not cells or not cells[0]:
                continue
            
            # Column 0: Date
            date_val = cells[0]
            if isinstance(date_val, datetime):
                d = date_val.date()
            elif isinstance(date_val, date):
                d = date_val
            elif isinstance(date_val, str):
                try:
                    d = datetime.strptime(date_val.strip(), "%d.%m.%Y").date()
                except ValueError:
                    continue
            else:
                continue
            
            # Prayer times: columns 2-7 (Fajr, Sunrise, Dhuhr, Asr, Maghrib, Isha)
            times = {}
            prayer_keys = ['fajr', 'sunrise', 'dhuhr', 'asr', 'maghrib', 'isha']
            for i, key in enumerate(prayer_keys):
                val = cells[2 + i] if (2 + i) < len(cells) else None
                if val:
                    if isinstance(val, datetime):
                        times[key] = val.strftime("%H:%M")
                    elif isinstance(val, str):
                        times[key] = val.strip()
                    else:
                        times[key] = str(val).strip()
            
            if times:
                daily_times[d] = times
            
            # Hijri date: last column (typically column 9 or later)
            # Look for the hijri date in the last few columns
            for col_idx in range(len(cells) - 1, 7, -1):
                val = cells[col_idx]
                if val and isinstance(val, str) and any(month in val for month in [
                    'Muharrem', 'Safer', 'Rebiülevvel', 'Rebiülahir',
                    'Cemaziyelevvel', 'Cemaziyelahir', 'Recep', 'Şaban',
                    'Ramazan', 'Şevval', 'Zilkade', 'Zilhicce',
                    # English hijri month names
                    'Muharram', 'Safar', 'Rabi', 'Jumada',
                    'Rajab', 'Shaban', 'Ramadan', 'Shawwal',
                    'Dhul', 'Dhu al'
                ]):
                    hijri_dates[d] = val.strip()
                    break
        
        wb.close()
        logger.info(f"Excel loaded: {len(daily_times)} days, city={city_name}")
        
    except Exception as e:
        logger.error(f"Excel load failed: {e}")
    
    return daily_times, hijri_dates, city_name


def load_prayer_data(config: dict, cache_path: str) -> tuple[dict, dict, str]:
    """Main entry point: load prayer data from best available source.
    
    Priority:
    1. If data_source == 'excel' and excel_path exists → load Excel
    2. If location_id is set → load from cache, fetch from API if cache expired
    3. Look for any .xlsx in the data directory as fallback
    
    Args:
        config: App configuration dict.
        cache_path: Path to the JSON cache file.
        
    Returns:
        Tuple of (daily_prayer_times, hijri_dates, city_name)
    """
    daily_times = {}
    hijri_dates = {}
    city_name = config.get('location_name', '') or "Prayer Times"
    
    data_source = config.get('data_source', 'api')
    
    # Option 1: Excel file
    if data_source == 'excel':
        excel_path = config.get('excel_path')
        if excel_path and os.path.exists(excel_path):
            daily_times, hijri_dates, city_name = load_excel_data(excel_path)
            if daily_times:
                return daily_times, hijri_dates, city_name
    
    # Option 2: API with cache
    location_id = config.get('location_id')
    if location_id:
        # Try cache first
        cache = load_cache(cache_path)
        if cache and cache.get('location_id') == location_id:
            cached_times = parse_api_data(cache.get('data', []))
            today = date.today()
            
            # Check if cache has today's data and at least 3 more days
            future_dates = [d for d in cached_times if d >= today]
            if len(future_dates) >= 3:
                logger.info(f"Using cached data: {len(future_dates)} future days")
                return cached_times, hijri_dates, city_name
            else:
                logger.info("Cache expired or insufficient, refreshing...")
        
        # Fetch fresh data from API
        raw_data = fetch_prayer_times(location_id)
        if raw_data:
            save_cache(cache_path, location_id, raw_data)
            daily_times = parse_api_data(raw_data)
            if daily_times:
                return daily_times, hijri_dates, city_name
    
    # Option 3: Fallback — look for any Excel in data directory
    data_dir = os.path.dirname(cache_path)
    if os.path.isdir(data_dir):
        for fname in os.listdir(data_dir):
            if fname.endswith('.xlsx'):
                fpath = os.path.join(data_dir, fname)
                daily_times, hijri_dates, city_name = load_excel_data(fpath)
                if daily_times:
                    return daily_times, hijri_dates, city_name
    
    # Also check project root for legacy Excel
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for fname in os.listdir(project_dir):
        if fname.endswith('.xlsx'):
            fpath = os.path.join(project_dir, fname)
            daily_times, hijri_dates, city_name = load_excel_data(fpath)
            if daily_times:
                return daily_times, hijri_dates, city_name
    
    return daily_times, hijri_dates, city_name


def refresh_if_needed(config: dict, cache_path: str, daily_times: dict) -> Optional[dict]:
    """Check if prayer data needs refreshing and fetch if so.
    
    Call this periodically (e.g. every hour) to keep data fresh.
    
    Returns:
        Updated daily_times dict if refreshed, None if no refresh needed.
    """
    if config.get('data_source') != 'api':
        return None
    
    location_id = config.get('location_id')
    if not location_id:
        return None
    
    today = date.today()
    future_dates = [d for d in daily_times if d >= today]
    
    # Refresh if less than 3 days of data remaining
    if len(future_dates) < 3:
        logger.info("Data running low, refreshing from API...")
        raw_data = fetch_prayer_times(location_id)
        if raw_data:
            save_cache(cache_path, location_id, raw_data)
            return parse_api_data(raw_data)
    
    return None

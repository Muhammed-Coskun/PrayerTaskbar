import sys
import os
import json
import ctypes
import winreg
from ctypes import wintypes
from datetime import datetime, timedelta
import openpyxl
from PyQt6.QtWidgets import (QApplication, QWidget, QLabel, QVBoxLayout, QHBoxLayout,
                             QGraphicsDropShadowEffect, QFrame)
from PyQt6.QtCore import Qt, QTimer, QPoint, QPropertyAnimation, QEasingCurve, QRect
from PyQt6.QtGui import QFont, QFontMetrics, QColor

# --- Thematische Icons pro Gebet ---
PRAYER_ICONS = {
    'İmsak':  '🌌',
    'Güneş':  '🌅',
    'Öğle':   '☀️',
    'İkindi': '🌤️',
    'Akşam':  '🌙',
    'Yatsı':  '🌕',
}

# --- Dini Günler (Diyanet 2026-2035) ---
# Format: (date(Y,M,D), "Name")
from datetime import date
HOLY_DAYS = [
    # ═══ 2026 ═══
    (date(2026,1,15), "Miraç Kandili"), (date(2026,2,2), "Berat Kandili"),
    (date(2026,2,19), "Ramazan Başlangıcı"), (date(2026,3,16), "Kadir Gecesi"),
    (date(2026,3,19), "Arefe"), (date(2026,3,20), "Ramazan Bayramı 1. Gün"),
    (date(2026,3,21), "Ramazan Bayramı 2. Gün"), (date(2026,3,22), "Ramazan Bayramı 3. Gün"),
    (date(2026,5,26), "Arefe"), (date(2026,5,27), "Kurban Bayramı 1. Gün"),
    (date(2026,5,28), "Kurban Bayramı 2. Gün"), (date(2026,5,29), "Kurban Bayramı 3. Gün"),
    (date(2026,5,30), "Kurban Bayramı 4. Gün"),
    (date(2026,6,16), "Hicri Yılbaşı"), (date(2026,6,25), "Aşure Günü"),
    (date(2026,8,24), "Mevlid Kandili"),
    (date(2026,12,10), "Üç Ayların Başlangıcı"), (date(2026,12,10), "Regaib Kandili"),
    # ═══ 2027 ═══
    (date(2027,1,4), "Miraç Kandili"), (date(2027,1,22), "Berat Kandili"),
    (date(2027,2,8), "Ramazan Başlangıcı"), (date(2027,3,5), "Kadir Gecesi"),
    (date(2027,3,8), "Arefe"), (date(2027,3,9), "Ramazan Bayramı 1. Gün"),
    (date(2027,3,10), "Ramazan Bayramı 2. Gün"), (date(2027,3,11), "Ramazan Bayramı 3. Gün"),
    (date(2027,5,15), "Arefe"), (date(2027,5,16), "Kurban Bayramı 1. Gün"),
    (date(2027,5,17), "Kurban Bayramı 2. Gün"), (date(2027,5,18), "Kurban Bayramı 3. Gün"),
    (date(2027,5,19), "Kurban Bayramı 4. Gün"),
    (date(2027,6,6), "Hicri Yılbaşı"), (date(2027,6,15), "Aşure Günü"),
    (date(2027,8,13), "Mevlid Kandili"),
    (date(2027,11,29), "Üç Ayların Başlangıcı"), (date(2027,12,2), "Regaib Kandili"),
    (date(2027,12,24), "Miraç Kandili"),
    # ═══ 2028 ═══
    (date(2028,1,11), "Berat Kandili"), (date(2028,1,28), "Ramazan Başlangıcı"),
    (date(2028,2,22), "Kadir Gecesi"), (date(2028,2,25), "Arefe"),
    (date(2028,2,26), "Ramazan Bayramı 1. Gün"), (date(2028,2,27), "Ramazan Bayramı 2. Gün"),
    (date(2028,2,28), "Ramazan Bayramı 3. Gün"),
    (date(2028,5,4), "Arefe"), (date(2028,5,5), "Kurban Bayramı 1. Gün"),
    (date(2028,5,6), "Kurban Bayramı 2. Gün"), (date(2028,5,7), "Kurban Bayramı 3. Gün"),
    (date(2028,5,8), "Kurban Bayramı 4. Gün"),
    (date(2028,5,25), "Hicri Yılbaşı"), (date(2028,6,3), "Aşure Günü"),
    (date(2028,8,2), "Mevlid Kandili"),
    (date(2028,11,18), "Üç Ayların Başlangıcı"), (date(2028,11,23), "Regaib Kandili"),
    (date(2028,12,13), "Miraç Kandili"), (date(2028,12,30), "Berat Kandili"),
    # ═══ 2029 ═══
    (date(2029,1,16), "Ramazan Başlangıcı"), (date(2029,2,10), "Kadir Gecesi"),
    (date(2029,2,13), "Arefe"), (date(2029,2,14), "Ramazan Bayramı 1. Gün"),
    (date(2029,2,15), "Ramazan Bayramı 2. Gün"), (date(2029,2,16), "Ramazan Bayramı 3. Gün"),
    (date(2029,4,23), "Arefe"), (date(2029,4,24), "Kurban Bayramı 1. Gün"),
    (date(2029,4,25), "Kurban Bayramı 2. Gün"), (date(2029,4,26), "Kurban Bayramı 3. Gün"),
    (date(2029,4,27), "Kurban Bayramı 4. Gün"),
    (date(2029,5,14), "Hicri Yılbaşı"), (date(2029,5,23), "Aşure Günü"),
    (date(2029,7,23), "Mevlid Kandili"),
    (date(2029,11,7), "Üç Ayların Başlangıcı"), (date(2029,11,8), "Regaib Kandili"),
    (date(2029,12,2), "Miraç Kandili"), (date(2029,12,20), "Berat Kandili"),
    # ═══ 2030 ═══
    (date(2030,1,5), "Ramazan Başlangıcı"), (date(2030,1,30), "Kadir Gecesi"),
    (date(2030,2,3), "Arefe"), (date(2030,2,4), "Ramazan Bayramı 1. Gün"),
    (date(2030,2,5), "Ramazan Bayramı 2. Gün"), (date(2030,2,6), "Ramazan Bayramı 3. Gün"),
    (date(2030,4,12), "Arefe"), (date(2030,4,13), "Kurban Bayramı 1. Gün"),
    (date(2030,4,14), "Kurban Bayramı 2. Gün"), (date(2030,4,15), "Kurban Bayramı 3. Gün"),
    (date(2030,4,16), "Kurban Bayramı 4. Gün"),
    (date(2030,5,3), "Hicri Yılbaşı"), (date(2030,5,12), "Aşure Günü"),
    (date(2030,7,12), "Mevlid Kandili"),
    (date(2030,10,28), "Üç Ayların Başlangıcı"), (date(2030,10,31), "Regaib Kandili"),
    (date(2030,11,22), "Miraç Kandili"), (date(2030,12,9), "Berat Kandili"),
    (date(2030,12,26), "Ramazan Başlangıcı"),
    # ═══ 2031 ═══
    (date(2031,1,20), "Kadir Gecesi"), (date(2031,1,23), "Arefe"),
    (date(2031,1,24), "Ramazan Bayramı 1. Gün"), (date(2031,1,25), "Ramazan Bayramı 2. Gün"),
    (date(2031,1,26), "Ramazan Bayramı 3. Gün"),
    (date(2031,4,1), "Arefe"), (date(2031,4,2), "Kurban Bayramı 1. Gün"),
    (date(2031,4,3), "Kurban Bayramı 2. Gün"), (date(2031,4,4), "Kurban Bayramı 3. Gün"),
    (date(2031,4,5), "Kurban Bayramı 4. Gün"),
    # ═══ 2032 ═══
    (date(2032,1,9), "Kadir Gecesi"), (date(2032,1,13), "Arefe"),
    (date(2032,1,14), "Ramazan Bayramı 1. Gün"), (date(2032,1,15), "Ramazan Bayramı 2. Gün"),
    (date(2032,1,16), "Ramazan Bayramı 3. Gün"),
    (date(2032,3,21), "Arefe"), (date(2032,3,22), "Kurban Bayramı 1. Gün"),
    (date(2032,3,23), "Kurban Bayramı 2. Gün"), (date(2032,3,24), "Kurban Bayramı 3. Gün"),
    (date(2032,3,25), "Kurban Bayramı 4. Gün"),
    # ═══ 2033 ═══
    (date(2033,1,1), "Arefe"), (date(2033,1,2), "Ramazan Bayramı 1. Gün"),
    (date(2033,1,3), "Ramazan Bayramı 2. Gün"), (date(2033,1,4), "Ramazan Bayramı 3. Gün"),
    (date(2033,3,10), "Arefe"), (date(2033,3,11), "Kurban Bayramı 1. Gün"),
    (date(2033,3,12), "Kurban Bayramı 2. Gün"), (date(2033,3,13), "Kurban Bayramı 3. Gün"),
    (date(2033,3,14), "Kurban Bayramı 4. Gün"),
    (date(2033,4,1), "Hicri Yılbaşı"), (date(2033,4,10), "Aşure Günü"),
    (date(2033,6,9), "Mevlid Kandili"),
    (date(2033,9,25), "Üç Ayların Başlangıcı"), (date(2033,9,29), "Regaib Kandili"),
    (date(2033,10,20), "Miraç Kandili"), (date(2033,11,6), "Berat Kandili"),
    (date(2033,11,23), "Ramazan Başlangıcı"), (date(2033,12,18), "Kadir Gecesi"),
    (date(2033,12,22), "Arefe"), (date(2033,12,23), "Ramazan Bayramı 1. Gün"),
    (date(2033,12,24), "Ramazan Bayramı 2. Gün"), (date(2033,12,25), "Ramazan Bayramı 3. Gün"),
    # ═══ 2034 ═══
    (date(2034,2,28), "Arefe"), (date(2034,3,1), "Kurban Bayramı 1. Gün"),
    (date(2034,3,2), "Kurban Bayramı 2. Gün"), (date(2034,3,3), "Kurban Bayramı 3. Gün"),
    (date(2034,3,4), "Kurban Bayramı 4. Gün"),
    (date(2034,3,21), "Hicri Yılbaşı"), (date(2034,3,30), "Aşure Günü"),
    (date(2034,5,29), "Mevlid Kandili"),
    (date(2034,9,14), "Üç Ayların Başlangıcı"), (date(2034,9,14), "Regaib Kandili"),
    (date(2034,10,9), "Miraç Kandili"), (date(2034,10,26), "Berat Kandili"),
    (date(2034,11,12), "Ramazan Başlangıcı"), (date(2034,12,7), "Kadir Gecesi"),
    (date(2034,12,11), "Arefe"), (date(2034,12,12), "Ramazan Bayramı 1. Gün"),
    (date(2034,12,13), "Ramazan Bayramı 2. Gün"), (date(2034,12,14), "Ramazan Bayramı 3. Gün"),
    # ═══ 2035 ═══
    (date(2035,2,17), "Arefe"), (date(2035,2,18), "Kurban Bayramı 1. Gün"),
    (date(2035,2,19), "Kurban Bayramı 2. Gün"), (date(2035,2,20), "Kurban Bayramı 3. Gün"),
    (date(2035,2,21), "Kurban Bayramı 4. Gün"),
    (date(2035,3,11), "Hicri Yılbaşı"), (date(2035,3,20), "Aşure Günü"),
    (date(2035,5,19), "Mevlid Kandili"),
    (date(2035,9,3), "Üç Ayların Başlangıcı"), (date(2035,9,6), "Regaib Kandili"),
    (date(2035,9,28), "Miraç Kandili"), (date(2035,10,16), "Berat Kandili"),
    (date(2035,11,1), "Ramazan Başlangıcı"), (date(2035,11,26), "Kadir Gecesi"),
    (date(2035,11,30), "Arefe"), (date(2035,12,1), "Ramazan Bayramı 1. Gün"),
    (date(2035,12,2), "Ramazan Bayramı 2. Gün"), (date(2035,12,3), "Ramazan Bayramı 3. Gün"),
]

# Visuelle Designs für jeden heiligen Tag-Typ (Gradient-basiert)
HOLY_DAY_STYLES = {
    'Miraç Kandili':         {'grad_start': '#1a0a2e', 'grad_mid': '#2d1b4e', 'grad_end': '#5c2d91', 'accent': '#b388ff', 'border': 'rgba(179, 136, 255, 0.30)'},
    'Berat Kandili':         {'grad_start': '#0d1525', 'grad_mid': '#1a2744', 'grad_end': '#2e3f6e', 'accent': '#90caf9', 'border': 'rgba(144, 202, 249, 0.25)'},
    'Kadir Gecesi':          {'grad_start': '#1a1400', 'grad_mid': '#2e2200', 'grad_end': '#5c4a00', 'accent': '#ffd700', 'border': 'rgba(255, 215, 0, 0.30)'},
    'Ramazan Başlangıcı':    {'grad_start': '#1a1500', 'grad_mid': '#2e2800', 'grad_end': '#4d4000', 'accent': '#ffb300', 'border': 'rgba(255, 179, 0, 0.25)'},
    'Ramazan Bayramı':       {'grad_start': '#0a1a0a', 'grad_mid': '#142e14', 'grad_end': '#1e5c1e', 'accent': '#66bb6a', 'border': 'rgba(102, 187, 106, 0.30)'},
    'Kurban Bayramı':        {'grad_start': '#0a1a18', 'grad_mid': '#142e2a', 'grad_end': '#1a5c52', 'accent': '#26a69a', 'border': 'rgba(38, 166, 154, 0.30)'},
    'Hicri Yılbaşı':         {'grad_start': '#0a1525', 'grad_mid': '#142844', 'grad_end': '#1e4a7a', 'accent': '#42a5f5', 'border': 'rgba(66, 165, 245, 0.30)'},
    'Aşure Günü':            {'grad_start': '#1a120a', 'grad_mid': '#2e1f12', 'grad_end': '#5c3a1a', 'accent': '#e6a44a', 'border': 'rgba(191, 128, 64, 0.25)'},
    'Mevlid Kandili':        {'grad_start': '#1a0a12', 'grad_mid': '#2e1420', 'grad_end': '#5c2840', 'accent': '#f48fb1', 'border': 'rgba(244, 143, 177, 0.30)'},
    'Regaib Kandili':        {'grad_start': '#0a1a1e', 'grad_mid': '#142e35', 'grad_end': '#1a5c68', 'accent': '#4dd0e1', 'border': 'rgba(77, 208, 225, 0.25)'},
    'Üç Ayların Başlangıcı': {'grad_start': '#140a25', 'grad_mid': '#251444', 'grad_end': '#3d1e7a', 'accent': '#b39ddb', 'border': 'rgba(126, 87, 194, 0.30)'},
}

def get_next_holy_day():
    """Findet den nächsten heiligen Tag ab heute (ohne Arefe)."""
    today = date.today()
    for d, name in sorted(HOLY_DAYS, key=lambda x: x[0]):
        if d >= today and name != 'Arefe':
            return d, name
    return None, None

user32 = ctypes.windll.user32

# --- Win32 API Fix ---
GWL_STYLE = -16
WS_CHILD = 0x40000000
WS_POPUP = 0x80000000

if sys.maxsize > 2**32:
    GetWindowLong = user32.GetWindowLongPtrW
    SetWindowLong = user32.SetWindowLongPtrW
else:
    GetWindowLong = user32.GetWindowLongW
    SetWindowLong = user32.SetWindowLongW

GetWindowLong.argtypes = [wintypes.HWND, ctypes.c_int]
GetWindowLong.restype = ctypes.c_void_p
SetWindowLong.argtypes = [wintypes.HWND, ctypes.c_int, ctypes.c_void_p]
SetWindowLong.restype = ctypes.c_void_p

# FindWindowExW für Multi-Monitor (c_void_p statt HWND für NULL-sichere Aufrufe)
user32.FindWindowExW.argtypes = [ctypes.c_void_p, ctypes.c_void_p, wintypes.LPCWSTR, wintypes.LPCWSTR]
user32.FindWindowExW.restype = ctypes.c_void_p

# IsWindow für HWND-Validierung (Monitor-Disconnect erkennen)
user32.IsWindow.argtypes = [ctypes.c_void_p]
user32.IsWindow.restype = ctypes.c_int

# --- Config-System ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")

def load_config():
    """Lädt die gespeicherte Konfiguration oder gibt Standardwerte zurück."""
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {"monitor_index": 0, "x_offset": 185}

def save_config(config):
    """Speichert die Konfiguration in config.json."""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2)
    except:
        pass

# --- Multi-Monitor Taskbar-Enumeration ---
def find_all_taskbars():
    """Findet alle Taskleisten: Shell_TrayWnd (primär) + alle Shell_SecondaryTrayWnd."""
    taskbars = []
    try:
        primary = user32.FindWindowW("Shell_TrayWnd", None)
        if primary:
            taskbars.append(("Ekran 1", int(primary)))
        
        prev_hwnd = 0
        idx = 2
        while True:
            hwnd = user32.FindWindowExW(0, prev_hwnd, "Shell_SecondaryTrayWnd", None)
            if not hwnd:
                break
            taskbars.append((f"Ekran {idx}", int(hwnd)))
            prev_hwnd = hwnd
            idx += 1
    except:
        pass
    
    return taskbars

# --- 1. Excel Daten laden ---
daily_prayer_times = {}
hijri_dates = {}  # Hicri Takvim
city_name = "Namaz Vakitleri"  # Fallback
is_ramadan = False  # Ramazan-Erkennung

def load_excel_data():
    global daily_prayer_times, hijri_dates, city_name, is_ramadan
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(script_dir, "Gebetszeiten.xlsx")
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active
        
        ay_cevirisi = {"Ocak": 1, "Şubat": 2, "Mart": 3, "Nisan": 4, "Mayıs": 5, "Haziran": 6, "Temmuz": 7, "Ağustos": 8, "Eylül": 9, "Ekim": 10, "Kasım": 11, "Aralık": 12}
        
        # Stadtname aus Zeile 2 extrahieren (z.B. "Weıl am rheın İçin Yıllık Namaz Vakti")
        raw_title = str(sheet.cell(row=2, column=1).value or "")
        if "İçin" in raw_title:
            city_name = raw_title.split("İçin")[0].strip().title()
        
        for row in range(1, sheet.max_row + 1):
            tarih_str = sheet.cell(row=row, column=1).value
            if not tarih_str: continue
            parts = str(tarih_str).split()
            if len(parts) >= 3 and parts[0].isdigit() and parts[1] in ay_cevirisi:
                date_obj = datetime(int(parts[2]), ay_cevirisi[parts[1]], int(parts[0])).date()
                daily_prayer_times[date_obj] = {
                    'İmsak': sheet.cell(row=row, column=3).value,
                    'Güneş': sheet.cell(row=row, column=4).value,
                    'Öğle':  sheet.cell(row=row, column=5).value,
                    'İkindi': sheet.cell(row=row, column=6).value,
                    'Akşam': sheet.cell(row=row, column=7).value,
                    'Yatsı': sheet.cell(row=row, column=8).value
                }
                # Hicri tarih (Spalte 2)
                hicri = sheet.cell(row=row, column=2).value
                if hicri:
                    hijri_dates[date_obj] = str(hicri).strip()
        return True
    except Exception as e:
        print(f"Hata: {e}")
        return False

# --- 2. Das Taskleisten-Widget ---
class PrayerWidget(QWidget):
    def __init__(self):
        super().__init__()
        
        self.config = load_config()
        self._flyout = None
        self._dragging = False
        self._drag_start_x = 0
        
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        # FIX: Kein Fokus bedeutet keine schwarzen Boxen mehr beim Klicken!
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        
        # CLEAN WINDOWS 11 DESIGN
        self.setStyleSheet("""
            QWidget#MainContainer {
                background-color: transparent;
                border: 1px solid transparent;
                border-radius: 6px;
            }
            QWidget#MainContainer:hover {
                background-color: rgba(255, 255, 255, 18); 
                border: 1px solid rgba(255, 255, 255, 12); 
            }
            QLabel {
                color: white;
                background-color: transparent; 
                border: none;
                padding: 0px;
            }
        """)
        
        self.container = QWidget(self)
        self.container.setObjectName("MainContainer")
        self.container.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(self.container)
        
        inner_layout = QVBoxLayout(self.container)
        inner_layout.setContentsMargins(10, 0, 10, 0)
        
        self.label = QLabel("Lade...")
        self.label.setFont(QFont("Segoe UI Semibold", 10))
        self.label.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        inner_layout.addWidget(self.label, alignment=Qt.AlignmentFlag.AlignCenter)
        
        # --- Multi-Monitor: Alle Taskleisten finden ---
        self.taskbars = find_all_taskbars()
        self.taskbar_hwnd = None
        self.current_monitor_index = 0
        
        # Gespeicherten Monitor verwenden (Fallback: erster verfügbarer)
        saved_idx = self.config.get("monitor_index", 0)
        if self.taskbars:
            idx = min(saved_idx, len(self.taskbars) - 1)
            self._do_inject(self.taskbars[idx][1])
            self.current_monitor_index = idx
        else:
            # Fallback: Original-Logik für StartAllBack-Kompatibilität
            self.taskbar_hwnd = user32.FindWindowW("Shell_SecondaryTrayWnd", None)
            if not self.taskbar_hwnd:
                self.taskbar_hwnd = user32.FindWindowW("Shell_TrayWnd", None)
            if self.taskbar_hwnd:
                self._do_inject(int(self.taskbar_hwnd))

        self.update_widget()
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_widget)
        self.timer.start(10000)  # Alle 10 Sek. — Widget zeigt nur HH:MM
        
        # Periodischer Taskbar-Rescan (spät startende oder getrennte Monitore)
        self._rescan_timer = QTimer(self)
        self._rescan_timer.timeout.connect(self._rescan_taskbars)
        self._rescan_timer.start(5000)  # Alle 5 Sekunden
    
    def _rescan_taskbars(self):
        """Sucht nach neuen/verlorenen Taskleisten und reagiert automatisch."""
        # Prüfen ob aktuelle Taskbar noch gültig ist
        if self.taskbar_hwnd and not user32.IsWindow(self.taskbar_hwnd):
            # Widget ist tot (WM_DESTROY) → komplett neu erstellen
            self._recreate_widget()
            return
        
        # Neue Taskbars suchen (Monitor wurde angeschlossen)
        new_taskbars = find_all_taskbars()
        if len(new_taskbars) != len(self.taskbars):
            self.taskbars = new_taskbars
            saved_idx = self.config.get("monitor_index", 0)
            if saved_idx < len(self.taskbars) and saved_idx != self.current_monitor_index:
                self._recreate_widget(target_monitor=saved_idx)
    
    def _recreate_widget(self, target_monitor=None):
        """Erstellt ein komplett neues Widget (einziger Weg nach WM_DESTROY)."""
        global _active_widget
        self.timer.stop()
        self._rescan_timer.stop()
        # Monitor-Index für das neue Widget setzen
        if target_monitor is not None:
            self.config["monitor_index"] = target_monitor
        else:
            self.config["monitor_index"] = 0  # Fallback: Primär
        save_config(self.config)
        # Neues Widget nach kurzer Verzögerung erstellen
        QTimer.singleShot(1000, self._do_recreate)
    
    def _do_recreate(self):
        global _active_widget
        try:
            new_widget = PrayerWidget()
            new_widget.show()
            _active_widget = new_widget
        except:
            pass
        self.hide()
        self.deleteLater()
    
    def _do_inject(self, taskbar_hwnd_int):
        """Kern-Injection: SetParent + WS_CHILD."""
        try:
            self.taskbar_hwnd = taskbar_hwnd_int
            my_hwnd = int(self.winId())
            user32.SetParent(my_hwnd, self.taskbar_hwnd)
            style = GetWindowLong(my_hwnd, GWL_STYLE)
            if style:
                style = (style | WS_CHILD) & ~WS_POPUP
                SetWindowLong(my_hwnd, GWL_STYLE, style)
        except:
            self.taskbar_hwnd = None
    
    def switch_monitor(self, index):
        """Wechselt das Widget auf einen anderen Bildschirm via Recreation."""
        if not self.taskbars or index >= len(self.taskbars) or index == self.current_monitor_index:
            return
        self._recreate_widget(target_monitor=index)
    
    def adjust_position(self, delta):
        """Verschiebt das Widget um delta Pixel (positiv = nach links)."""
        x_offset = self.config.get("x_offset", 185)
        x_offset = max(10, x_offset + delta)
        self.config["x_offset"] = x_offset
        save_config(self.config)
        self.update_widget()

    def update_widget(self):
        # Prüfe ob Taskbar noch existiert
        if self.taskbar_hwnd and not user32.IsWindow(self.taskbar_hwnd):
            self._recreate_widget()
            return
        
        now = datetime.now()
        today = now.date()
        
        if today not in daily_prayer_times:
            self.label.setText("Datum fehlt!")
            return
            
        times_today = daily_prayer_times[today]
        next_prayer, next_prayer_time = None, None
        
        prayer_order = ['İmsak', 'Güneş', 'Öğle', 'İkindi', 'Akşam', 'Yatsı']
        
        for p_name in prayer_order:
            if p_name not in times_today: continue
            p_time = times_today[p_name]
            pt = datetime.strptime(p_time.strip(), "%H:%M").replace(year=now.year, month=now.month, day=now.day) if isinstance(p_time, str) else datetime.combine(today, p_time)
            
            if pt > now:
                next_prayer, next_prayer_time = p_name, pt
                break
                
        if not next_prayer:
            tomorrow = today + timedelta(days=1)
            if tomorrow in daily_prayer_times:
                time_val = daily_prayer_times[tomorrow]['İmsak']
                pt = datetime.strptime(time_val.strip(), "%H:%M").replace(year=tomorrow.year, month=tomorrow.month, day=tomorrow.day) if isinstance(time_val, str) else datetime.combine(tomorrow, time_val)
                next_prayer, next_prayer_time = 'İmsak', pt

        if next_prayer and next_prayer_time:
            diff = next_prayer_time - now
            
            # Farbe ändern bei < 15 Min
            if diff.total_seconds() <= 900:
                self.label.setStyleSheet("color: #ff8b8b; background-color: transparent; border: none; padding: 0px;")
            else:
                self.label.setStyleSheet("color: #e5e5e5; background-color: transparent; border: none; padding: 0px;")
                
            hours, remainder = divmod(diff.seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            
            # Cleanes Design OHNE Icon (Icons nur im Menü)
            display_text = f"{next_prayer}   {hours:02d}:{minutes:02d}"
            self.label.setText(display_text)

            # --- DYNAMISCHE BREITE BERECHNEN (Wrapper) ---
            if self.taskbar_hwnd:
                rect = wintypes.RECT()
                user32.GetClientRect(self.taskbar_hwnd, ctypes.byref(rect))
                tb_w = rect.right - rect.left
                tb_h = rect.bottom - rect.top
                
                if tb_h > 0:
                    # Berechnet die exakte Pixel-Breite des Textes
                    fm = QFontMetrics(self.label.font())
                    text_width = fm.horizontalAdvance(display_text)
                    
                    w = text_width + 24
                    h = tb_h - 6 
                    
                    x_offset = self.config.get("x_offset", 185)
                    x = tb_w - w - x_offset
                    y = (tb_h - h) // 2
                    
                    self.setGeometry(x, y, w, h)

    # --- Maus-Events ---
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.MiddleButton:
            self._dragging = True
            self._drag_start_x = event.globalPosition().x()
        elif event.button() in (Qt.MouseButton.LeftButton, Qt.MouseButton.RightButton):
            self.show_modern_menu()
    
    def mouseMoveEvent(self, event):
        if self._dragging:
            current_x = event.globalPosition().x()
            dx = int(self._drag_start_x - current_x)
            self._drag_start_x = current_x
            self.adjust_position(dx)
    
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.MouseButton.MiddleButton:
            self._dragging = False

    # --- 3. Das moderne Win11 Flyout-Menü ---
    def show_modern_menu(self):
        if hasattr(self, '_flyout') and self._flyout and self._flyout.isVisible():
            self._flyout.close()
            self._flyout = None
            return
        
        # Altes Flyout aufräumen (Memory-Leak verhindern)
        if hasattr(self, '_flyout') and self._flyout:
            self._flyout.deleteLater()
            self._flyout = None

        # Taskbar-Liste aktualisieren (falls Monitor inzwischen an ist)
        self.taskbars = find_all_taskbars()
        
        self._flyout = PrayerFlyout(self)
        
        rect = wintypes.RECT()
        user32.GetWindowRect(int(self.winId()), ctypes.byref(rect))
        flyout_w = self._flyout.sizeHint().width()
        flyout_h = self._flyout.sizeHint().height()
        
        menu_x = rect.left + (rect.right - rect.left) // 2 - flyout_w // 2
        menu_y = rect.top - flyout_h - 12
        
        # Slide-Up Animation: Startet 18px tiefer und gleitet sanft nach oben
        start_y = menu_y + 18
        self._flyout.setGeometry(menu_x, start_y, flyout_w, flyout_h)
        self._flyout.setWindowOpacity(0.0)
        self._flyout.show()
        self._flyout.activateWindow()
        
        # Positions-Animation (sanftes Gleiten)
        self._slide_anim = QPropertyAnimation(self._flyout, b"geometry")
        self._slide_anim.setDuration(280)
        self._slide_anim.setStartValue(QRect(menu_x, start_y, flyout_w, flyout_h))
        self._slide_anim.setEndValue(QRect(menu_x, menu_y, flyout_w, flyout_h))
        self._slide_anim.setEasingCurve(QEasingCurve.Type.OutQuad)
        self._slide_anim.start()
        
        # Opacity-Animation (weiches Einblenden)
        self._opacity_anim = QPropertyAnimation(self._flyout, b"windowOpacity")
        self._opacity_anim.setDuration(250)
        self._opacity_anim.setStartValue(0.0)
        self._opacity_anim.setEndValue(1.0)
        self._opacity_anim.setEasingCurve(QEasingCurve.Type.OutQuad)
        self._opacity_anim.start()


# --- 4. Win11-Style Flyout Panel ---
class PrayerFlyout(QWidget):
    """Custom Win11 Quick-Settings-artiges Flyout für Gebetszeiten."""
    
    def __init__(self, prayer_widget):
        super().__init__(None)  # Kein Parent → eigenes Top-Level-Fenster
        self.prayer_widget = prayer_widget
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool |
            Qt.WindowType.Popup
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedWidth(340)
        
        # Drop Shadow wie Win11
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(32)
        shadow.setColor(QColor(0, 0, 0, 120))
        shadow.setOffset(0, 4)
        self.setGraphicsEffect(shadow)
        
        self._build_ui()
        
        # Auto-Close: Schließt sich wenn das Fenster nicht mehr aktiv ist
        self._close_timer = QTimer(self)
        self._close_timer.timeout.connect(self._check_active)
        self._close_timer.start(300)
    
    def _check_active(self):
        if not self.isActiveWindow():
            self._close_timer.stop()
            self.close()
    
    def _build_ui(self):
        container = QWidget(self)
        container.setObjectName("FlyoutContainer")
        container.setStyleSheet("""
            QWidget#FlyoutContainer {
                background-color: #2b2b2b;
                border: 1px solid rgba(255, 255, 255, 0.08);
                border-radius: 12px;
            }
        """)
        
        outer = QVBoxLayout(self)
        outer.setContentsMargins(8, 8, 8, 8)
        outer.addWidget(container)
        
        layout = QVBoxLayout(container)
        layout.setContentsMargins(18, 18, 18, 14)
        layout.setSpacing(5)
        
        # ═══════ HEADER ═══════
        header_label = QLabel(city_name)
        header_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI Variable Display'; font-size: 15pt; font-weight: 600; background: transparent; border: none;")
        layout.addWidget(header_label)
        
        # Tarih satırı: Hicri | Miladi yan yana
        today = datetime.now().date()
        now = datetime.now()
        hicri_str = hijri_dates.get(today, "")
        
        gun_isimleri = {0: 'Pazartesi', 1: 'Salı', 2: 'Çarşamba', 3: 'Perşembe', 4: 'Cuma', 5: 'Cumartesi', 6: 'Pazar'}
        ay_isimleri = {1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran', 7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'}
        miladi_str = f"{gun_isimleri[now.weekday()]}, {now.day} {ay_isimleri[now.month]} {now.year}"

        date_row = QHBoxLayout()
        date_row.setSpacing(0)
        date_row.setContentsMargins(0, 0, 0, 0)
        
        miladi_label = QLabel(miladi_str)
        miladi_label.setStyleSheet("color: #8a8a8a; font-family: 'Segoe UI'; font-size: 9pt; background: transparent; border: none;")
        date_row.addWidget(miladi_label)
        
        if hicri_str:
            dot = QLabel("  ·  ")
            dot.setStyleSheet("color: #555555; font-family: 'Segoe UI'; font-size: 9pt; background: transparent; border: none;")
            date_row.addWidget(dot)
            hicri_label = QLabel(hicri_str)
            hicri_label.setStyleSheet("color: #8a8a8a; font-family: 'Segoe UI'; font-size: 9pt; background: transparent; border: none;")
            date_row.addWidget(hicri_label)
        
        date_row.addStretch()
        layout.addLayout(date_row)
        
        layout.addSpacing(8)
        layout.addWidget(self._make_separator())
        layout.addSpacing(4)
        
        # ═══════ DİNİ GÜNLER ═══════
        holy_date, holy_name = get_next_holy_day()
        if holy_date and holy_name:
            layout.addWidget(self._create_holy_day_card(holy_date, holy_name))
            layout.addSpacing(4)
            layout.addWidget(self._make_separator())
            layout.addSpacing(4)
        
        # ═══════ GEBETSZEITEN ═══════
        now_time = datetime.now()
        
        if today in daily_prayer_times:
            next_p = None
            for p_name in ['İmsak', 'Güneş', 'Öğle', 'İkindi', 'Akşam', 'Yatsı']:
                if p_name in daily_prayer_times[today]:
                    p_time = daily_prayer_times[today][p_name]
                    pt = datetime.strptime(p_time.strip(), "%H:%M").replace(year=now_time.year, month=now_time.month, day=now_time.day) if isinstance(p_time, str) else datetime.combine(today, p_time)
                    if pt > now_time:
                        next_p = p_name
                        break
            
            for p_name in ['İmsak', 'Güneş', 'Öğle', 'İkindi', 'Akşam', 'Yatsı']:
                if p_name in daily_prayer_times[today]:
                    p_time = daily_prayer_times[today][p_name]
                    pt = datetime.strptime(p_time.strip(), "%H:%M").replace(year=now_time.year, month=now_time.month, day=now_time.day) if isinstance(p_time, str) else datetime.combine(today, p_time)
                    time_str = p_time.strip() if isinstance(p_time, str) else p_time.strftime("%H:%M")
                    icon = PRAYER_ICONS.get(p_name, '')
                    is_current = (p_name == next_p)
                    is_past = (pt < now_time)
                    # Ramadan Iftar: Akşam bekommt spezielles Styling
                    is_iftar = (p_name == 'Akşam' and is_ramadan)
                    layout.addWidget(self._create_prayer_row(icon, p_name, time_str, is_current, is_past, is_iftar))
        
        layout.addSpacing(6)
        layout.addWidget(self._make_separator())
        
        # ═══════ FOOTER: Ekran + Konum + Kapat ═══════
        footer = QHBoxLayout()
        footer.setSpacing(5)
        footer.setContentsMargins(0, 4, 0, 0)
        
        # Monitor-Buttons (nur wenn >1)
        if len(self.prayer_widget.taskbars) > 1:
            for i, (name, hwnd) in enumerate(self.prayer_widget.taskbars):
                is_active = (i == self.prayer_widget.current_monitor_index)
                btn = self._make_monitor_btn(name, i, is_active)
                footer.addWidget(btn)
            
            # Dezenter vertikaler Trenner
            sep = QLabel("│")
            sep.setStyleSheet("color: #3d3d3d; font-size: 12pt; background: transparent; border: none;")
            sep.setFixedWidth(14)
            footer.addWidget(sep)
        
        # Position-Buttons
        left_btn = self._make_icon_btn("◂")
        left_btn.mousePressEvent = lambda e: self.prayer_widget.adjust_position(20)
        footer.addWidget(left_btn)
        
        right_btn = self._make_icon_btn("▸")
        right_btn.mousePressEvent = lambda e: self.prayer_widget.adjust_position(-20)
        footer.addWidget(right_btn)
        
        footer.addStretch()
        
        # Autostart-Button (nur wenn noch nicht aktiviert)
        if not self._is_autostart_enabled():
            auto_btn = self._make_small_btn("Otomatik")
            auto_btn.mousePressEvent = lambda e: self._toggle_autostart()
            footer.addWidget(auto_btn)
        
        # Kapat
        exit_btn = self._make_small_btn("Kapat")
        exit_btn.mousePressEvent = lambda e: QApplication.instance().quit()
        footer.addWidget(exit_btn)
        
        layout.addLayout(footer)
    
    def _make_separator(self):
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: rgba(255, 255, 255, 0.08); border: none; max-height: 1px;")
        return line
    
    def _create_holy_day_card(self, holy_date, holy_name):
        """Erstellt eine visuell einzigartige Karte mit Gradient-Hintergrund — Text links."""
        # Style-Key ermitteln
        style_key = holy_name
        for base in ['Ramazan Bayramı', 'Kurban Bayramı']:
            if holy_name.startswith(base):
                style_key = base
                break
        
        style = HOLY_DAY_STYLES.get(style_key, HOLY_DAY_STYLES.get('Kadir Gecesi'))
        
        card = QWidget()
        card.setFixedHeight(100)
        
        card.setStyleSheet(f"""
            QWidget {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0.3,
                    stop:0 {style['grad_start']},
                    stop:0.45 {style['grad_mid']},
                    stop:1.0 {style['grad_end']});
                border: none;
                border-radius: 10px;
            }}
        """)
        
        v_layout = QVBoxLayout(card)
        v_layout.setContentsMargins(16, 10, 14, 10)
        v_layout.setSpacing(1)
        
        # Zeile 1: Countdown — große Zahl + "Gün"
        delta = (holy_date - date.today()).days
        countdown_row = QHBoxLayout()
        countdown_row.setSpacing(4)
        countdown_row.setContentsMargins(0, 0, 0, 0)
        
        if delta == 0:
            num_label = QLabel("✦")
            num_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI Variable Display'; font-size: 22pt; font-weight: 700; background: transparent; border: none;")
            countdown_row.addWidget(num_label)
            unit_label = QLabel("Bugün!")
            unit_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI'; font-size: 11pt; font-weight: 600; background: transparent; border: none;")
            countdown_row.addWidget(unit_label, alignment=Qt.AlignmentFlag.AlignBottom)
        elif delta == 1:
            num_label = QLabel("Yarın")
            num_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI Variable Display'; font-size: 22pt; font-weight: 700; background: transparent; border: none;")
            countdown_row.addWidget(num_label)
        else:
            num_label = QLabel(str(delta))
            num_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI Variable Display'; font-size: 22pt; font-weight: 700; background: transparent; border: none;")
            countdown_row.addWidget(num_label)
            unit_label = QLabel("Gün")
            unit_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI'; font-size: 10pt; font-weight: 800; background: transparent; border: none; padding-top: 12px;")
            countdown_row.addWidget(unit_label, alignment=Qt.AlignmentFlag.AlignVCenter)
        
        countdown_row.addStretch()
        v_layout.addLayout(countdown_row)
        
        # Zeile 2: Name des heiligen Tages (fett)
        name_label = QLabel(holy_name)
        name_label.setStyleSheet("color: #ffffff; font-family: 'Segoe UI Variable'; font-size: 11pt; font-weight: 700; background: transparent; border: none;")
        name_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        v_layout.addWidget(name_label)
        
        # Zeile 3: Miladi + Hicri Datum
        ay_isimleri = {1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran', 7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık'}
        date_str = f"{holy_date.day} {ay_isimleri[holy_date.month]} {holy_date.year}"
        
        # Hicri-Datum aus der globalen hijri_dates holen
        hicri_str = hijri_dates.get(holy_date, "")
        if hicri_str:
            info_text = f"{date_str}  │  {hicri_str}"
        else:
            info_text = date_str
        
        info_label = QLabel(info_text)
        info_label.setStyleSheet("color: rgba(255,255,255,0.55); font-family: 'Segoe UI'; font-size: 8pt; background: transparent; border: none;")
        info_label.setAlignment(Qt.AlignmentFlag.AlignLeft)
        v_layout.addWidget(info_label)
        
        return card
    
    def _make_small_btn(self, text):
        btn = QLabel(text)
        btn.setStyleSheet("""
            color: #8a8a8a;
            font-family: 'Segoe UI';
            font-size: 8pt;
            background-color: rgba(255, 255, 255, 0.05);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 5px;
            padding: 4px 10px;
        """)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        return btn
    
    def _make_icon_btn(self, text):
        """Kompakter Button nur für Icons (Pfeile etc.)"""
        btn = QLabel(text)
        btn.setStyleSheet("""
            color: #8a8a8a;
            font-family: 'Segoe UI';
            font-size: 9pt;
            background-color: rgba(255, 255, 255, 0.05);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 5px;
            padding: 3px 8px;
        """)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        return btn
    
    def _make_monitor_btn(self, name, index, is_active):
        btn = QLabel(name)
        if is_active:
            btn.setStyleSheet("""
                color: #60cdff; font-family: 'Segoe UI'; font-size: 8pt; font-weight: 600;
                background-color: rgba(96, 205, 255, 0.12);
                border: 1px solid rgba(96, 205, 255, 0.25);
                border-radius: 4px; padding: 3px 8px;
            """)
        else:
            btn.setStyleSheet("""
                color: #e5e5e5; font-family: 'Segoe UI'; font-size: 8pt;
                background-color: rgba(255, 255, 255, 0.06);
                border: 1px solid rgba(255, 255, 255, 0.08);
                border-radius: 4px; padding: 3px 8px;
            """)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.mousePressEvent = lambda e, idx=index: self._on_monitor_click(idx)
        return btn
    
    def _on_monitor_click(self, index):
        self.prayer_widget.switch_monitor(index)
        self.close()
    
    def _create_prayer_row(self, icon, name, time_str, is_current, is_past, is_iftar=False):
        row = QWidget()
        row.setFixedHeight(40)
        
        if is_iftar and not is_past:
            # ✨ RAMADAN IFTAR - Goldenes Styling mit Dekoration
            row.setStyleSheet("""
                QWidget {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 rgba(255, 180, 50, 0.15),
                        stop:0.5 rgba(255, 200, 80, 0.20),
                        stop:1 rgba(255, 180, 50, 0.15));
                    border: 1px solid rgba(255, 190, 60, 0.35);
                    border-radius: 8px;
                }
            """)
            name_color = "#ffc850"
            time_color = "#ffc850"
            font_weight = "600"
            # Iftar-Name mit Dekoration
            name = "✦ İftar ✦"
            icon = "🌙"
        elif is_current:
            row.setStyleSheet("""
                QWidget {
                    background-color: rgba(96, 205, 255, 0.12);
                    border: 1px solid rgba(96, 205, 255, 0.25);
                    border-radius: 8px;
                }
            """)
            name_color = "#60cdff"
            time_color = "#60cdff"
            font_weight = "600"
        elif is_past:
            row.setStyleSheet("""
                QWidget {
                    background-color: transparent;
                    border: 1px solid transparent;
                    border-radius: 8px;
                }
            """)
            name_color = "#7a7a7a"
            time_color = "#6a6a6a"
            font_weight = "400"
        else:
            row.setStyleSheet("""
                QWidget {
                    background-color: rgba(255, 255, 255, 0.04);
                    border: 1px solid rgba(255, 255, 255, 0.06);
                    border-radius: 8px;
                }
            """)
            name_color = "#e5e5e5"
            time_color = "#b0b0b0"
            font_weight = "400"
        
        h_layout = QHBoxLayout(row)
        h_layout.setContentsMargins(12, 0, 14, 0)
        h_layout.setSpacing(0)
        
        icon_label = QLabel(icon)
        icon_label.setFixedWidth(28)
        icon_label.setStyleSheet("font-size: 14pt; background: transparent; border: none;")
        h_layout.addWidget(icon_label)
        
        name_label = QLabel(name)
        name_label.setStyleSheet(f"color: {name_color}; font-family: 'Segoe UI Variable'; font-size: 10pt; font-weight: {font_weight}; background: transparent; border: none;")
        h_layout.addWidget(name_label)
        
        h_layout.addStretch()
        
        time_label = QLabel(time_str)
        time_label.setStyleSheet(f"color: {time_color}; font-family: 'Segoe UI'; font-size: 10pt; font-weight: {font_weight}; background: transparent; border: none;")
        time_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        h_layout.addWidget(time_label)
        
        return row
    
    def _is_autostart_enabled(self):
        """Prüft ob das Programm im Windows-Autostart ist."""
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_READ)
            winreg.QueryValueEx(key, "GebetszeitenWidget")
            winreg.CloseKey(key)
            return True
        except:
            return False
    
    def _toggle_autostart(self):
        """Schaltet den Windows-Autostart ein/aus."""
        try:
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Run", 0, winreg.KEY_SET_VALUE | winreg.KEY_READ)
            if self._is_autostart_enabled():
                winreg.DeleteValue(key, "GebetszeitenWidget")
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                script_path = os.path.abspath(__file__)
                # pythonw.exe im selben Verzeichnis wie python.exe suchen (kein CMD-Fenster)
                python_dir = os.path.dirname(sys.executable)
                pythonw = os.path.join(python_dir, "pythonw.exe")
                if not os.path.exists(pythonw):
                    pythonw = sys.executable  # Fallback
                winreg.SetValueEx(key, "GebetszeitenWidget", 0, winreg.REG_SZ, f'"{pythonw}" "{script_path}"')
            winreg.CloseKey(key)
        except:
            pass
        self.close()
    
    def sizeHint(self):
        return self.layout().sizeHint() if self.layout() else super().sizeHint()
    
    def focusOutEvent(self, event):
        self.close()
        super().focusOutEvent(event)

_active_widget = None

if __name__ == "__main__":
    # Stderr auf /dev/null umleiten (kein wachsender Puffer im RAM)
    sys.stderr = open(os.devnull, 'w')
    
    app = QApplication(sys.argv)
    if not load_excel_data():
        sys.exit()
    
    # Ramazan-Erkennung für heute
    today = datetime.now().date()
    hicri_today = hijri_dates.get(today, "")
    if "Ramazan" in hicri_today:
        is_ramadan = True
    
    _active_widget = PrayerWidget()
    _active_widget.show()
    sys.exit(app.exec())